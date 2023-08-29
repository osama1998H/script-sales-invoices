import os
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm


FILE_PATH = "output/"
FILES = os.listdir(FILE_PATH)


template_name = "heading_template.xlsx"


template_df = pd.read_excel(template_name, header=0)

freq = {column: template_df[column][0] for column in template_df.columns}
for excel_file in tqdm(
    FILES, desc="Clear Data |+| Calculate Total Values:  ", colour="GREEN", unit=" File"
):

    main_df = pd.read_excel(FILE_PATH + excel_file, header=0)

    for column in main_df.columns:
        for key, value in freq.items():
            if column == key:
                if "item" in str(key).lower():
                    # print("pass")
                    pass

                    # ! write code here
                    # * we passed the condition here
                elif value == 1:
                    # ! write code here
                    # ? loop for delete cells in data

                    for i, item in enumerate(main_df[column]):
                        if i != 0:
                            main_df.at[i, column] = None

    main_df.to_excel(FILE_PATH + excel_file, header=True, index=False)

    # todo // open file with openpyxl
    wb_object = load_workbook(FILE_PATH + excel_file)

    # todo // select active sheet
    sheet_object = wb_object.active

    # todo // add SUM function to the cell

    max_col = sheet_object.max_column
    max_row = sheet_object.max_row
    amount_cord = None
    quantity_cord = None
    total_quantity_cord = None
    total_cord = None
    # Loop will print all columns name
    for i in range(1, max_col + 1):

        cell_obj = sheet_object.cell(row=1, column=i)
        if cell_obj.value == "Amount (Items)":
            amount_cord = cell_obj.coordinate

        elif cell_obj.value == "Quantity (Items)":
            quantity_cord = cell_obj.coordinate

        elif cell_obj.value == "Total Quantity":
            total_quantity_cord = cell_obj.coordinate

        elif cell_obj.value == "Total":
            total_cord = cell_obj.coordinate

    net_amount_cord = "".join([i for i in amount_cord if not i.isdigit()])
    start_amount_cord = f"{net_amount_cord}2"
    end_amount_cord = net_amount_cord + str(max_row)

    net_quantity_cord = "".join([i for i in quantity_cord if not i.isdigit()])
    start_quantity_cord = f"{net_quantity_cord}2"
    end_quantity_cord = net_quantity_cord + str(max_row)

    #  todo
    net_total_quantity_cord = "".join(
        [i for i in total_quantity_cord if not i.isdigit()]
    )
    start_total_quantity_cord = f"{net_total_quantity_cord}2"
    end_total_quantity_cord = net_total_quantity_cord + str(max_row)
    #  todo

    net_total_cord = "".join([i for i in total_cord if not i.isdigit()])
    start_total_cord = f"{net_total_cord}2"
    end_total_cord = net_total_cord + str(max_row)

    # sheet_object[start_total_quantity_cord] = f"=SUM({start_quantity_cord}:{end_quantity_cord})"
    # sheet_object[start_total_cord] = f"=SUM({start_amount_cord}:{end_amount_cord})"

    amount_data = [net_amount_cord + str(i + 1) for i in range(1, max_row + 1)]
    quantity_data = [net_quantity_cord + str(i + 1) for i in range(1, max_row + 1)]
    total = 0
    total_quantity = 0
    amount_data.pop()
    quantity_data.pop()

    for amount_value, total_quantity_value in zip(amount_data, quantity_data):

        total += sheet_object[amount_value].value
        total_quantity += sheet_object[total_quantity_value].value

    sheet_object[start_total_quantity_cord] = total_quantity
    sheet_object[start_total_cord] = total

    # todo // close the sheet & file

    wb_object.save(FILE_PATH + excel_file)
